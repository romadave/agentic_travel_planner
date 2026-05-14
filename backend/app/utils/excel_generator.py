import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from app.models.final_trip_request import ItineraryOption, FlightOption


def generate_itinerary_excel(itinerary: ItineraryOption, flights: list[FlightOption]) -> bytes:
    wb = Workbook()
    _build_itinerary_sheet(wb.active, itinerary)
    _build_flights_sheet(wb.create_sheet("Flights"), flights)
    _build_hotels_sheet(wb.create_sheet("Hotels"), itinerary)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_itinerary_sheet(ws, itinerary: ItineraryOption):
    ws.title = "Itinerary"
    headers = [
        "Day", "Date", "Area",
        "Morning Activity", "Morning Place", "Morning Food",
        "Afternoon Activity", "Afternoon Place", "Afternoon Food",
        "Evening Activity", "Evening Place", "Evening Food",
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    for day in itinerary.days:
        ws.append([
            day.dayNumber,
            day.date,
            day.area,
            day.morning.activity if day.morning else "",
            day.morning.place if day.morning else "",
            day.morning.foodSuggestion if day.morning else "",
            day.afternoon.activity if day.afternoon else "",
            day.afternoon.place if day.afternoon else "",
            day.afternoon.foodSuggestion if day.afternoon else "",
            day.evening.activity if day.evening else "",
            day.evening.place if day.evening else "",
            day.evening.foodSuggestion if day.evening else "",
        ])
    _auto_width(ws)


def _build_flights_sheet(ws, flights: list[FlightOption]):
    ws.title = "Flights"
    headers = [
        "Rank", "Airline", "Flight #", "Origin", "Destination",
        "Departure Date", "Departure Time", "Return Date", "Return Time",
        "Duration (hrs)", "Layovers", "Price ($)", "Booking URL",
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    for f in sorted(flights, key=lambda x: x.rank):
        ws.append([
            f.rank, f.airline, f.flightNumber or "",
            f.origin, f.destination,
            f.departureDate, f.departureTime,
            f.returnDate, f.returnTime,
            f.duration, ", ".join(f.layovers),
            f.price, f.bookingUrl or "",
        ])
    _auto_width(ws)


def _build_hotels_sheet(ws, itinerary: ItineraryOption):
    ws.title = "Hotels"
    headers = [
        "Area", "Nights", "Hotel Name", "Type", "Rating",
        "Price/Night ($)", "Total Price ($)", "Amenities", "Booking URL",
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    for stop in itinerary.hotelStops:
        for hotel in stop.hotels:
            ws.append([
                stop.area, stop.nights,
                hotel.name, hotel.type or "", hotel.rating,
                hotel.pricePerNight, hotel.totalPrice,
                ", ".join(hotel.amenities), hotel.bookingUrl or "",
            ])
    _auto_width(ws)


def _style_header_row(ws, row: int, col_count: int):
    fill = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
